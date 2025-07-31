"""
Reporting Service for RMS
Handles CSV and Excel report generation without pandas dependency
"""

import io
import csv
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.models import User
from django.db.models import Count, Avg, Q
from .models import (
    Student, Result, Course, Faculty, Department, Level,
    AcademicSession, CourseEnrollment, CarryOverList
)

try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ReportingService:
    """Service class for generating reports and exports"""

    @staticmethod
    def generate_csv_response(data, filename):
        """
        Generate CSV file response

        Args:
            data: List of dictionaries
            filename: Name of the file

        Returns:
            HttpResponse with CSV file
        """
        if not data:
            # Return empty CSV if no data
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            return response

        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        # Get field names from first row
        fieldnames = data[0].keys()
        writer = csv.DictWriter(response, fieldnames=fieldnames)

        # Write header and data
        writer.writeheader()
        for row in data:
            writer.writerow(row)

        return response

    @staticmethod
    def generate_excel_response(data, filename, sheet_name="Sheet1"):
        """
        Generate Excel file response (fallback to CSV if openpyxl not available)

        Args:
            data: List of dictionaries
            filename: Name of the file
            sheet_name: Name of the Excel sheet

        Returns:
            HttpResponse with Excel or CSV file
        """
        if not EXCEL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            return ReportingService.generate_csv_response(data, filename)

        if not data:
            # Return empty Excel if no data
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Get field names from first row
        fieldnames = list(data[0].keys())

        # Write header
        for col_num, header in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, field in enumerate(fieldnames, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(field, ''))

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
    
    @staticmethod
    def get_faculty_summary_data(faculty=None):
        """
        Get faculty summary data for reporting
        
        Args:
            faculty: Faculty object (optional, if None returns all faculties)
        
        Returns:
            Dictionary with faculty statistics
        """
        if faculty:
            faculties = [faculty]
        else:
            faculties = Faculty.objects.all()
        
        summary_data = []
        
        for fac in faculties:
            departments = Department.objects.filter(faculty=fac)
            students = Student.objects.filter(faculty=fac)
            courses = Course.objects.filter(departments__faculty=fac).distinct()
            
            # Calculate statistics
            total_students = students.count()
            total_departments = departments.count()
            total_courses = courses.count()
            
            # Results statistics
            published_results = Result.objects.filter(
                enrollment__student__faculty=fac,
                status='PUBLISHED'
            )
            
            carryovers = published_results.filter(is_carry_over=True).count()
            pass_rate = 0
            if published_results.count() > 0:
                pass_rate = ((published_results.count() - carryovers) / published_results.count()) * 100
            
            summary_data.append({
                'Faculty': fac.name,
                'Departments': total_departments,
                'Students': total_students,
                'Courses': total_courses,
                'Published Results': published_results.count(),
                'Carryovers': carryovers,
                'Pass Rate (%)': round(pass_rate, 2)
            })
        
        return summary_data
    
    @staticmethod
    def get_student_results_data(student=None, level=None, faculty=None, session=None):
        """
        Get student results data for reporting
        
        Args:
            student: Student object (optional)
            level: Level object (optional)
            faculty: Faculty object (optional)
            session: AcademicSession object (optional)
        
        Returns:
            List of dictionaries with student results
        """
        results = Result.objects.filter(status='PUBLISHED').select_related(
            'enrollment__student',
            'enrollment__course',
            'enrollment__session'
        )
        
        # Apply filters
        if student:
            results = results.filter(enrollment__student=student)
        if level:
            results = results.filter(enrollment__student__current_level=level)
        if faculty:
            results = results.filter(enrollment__student__faculty=faculty)
        if session:
            results = results.filter(enrollment__session=session)
        
        results_data = []
        for result in results:
            results_data.append({
                'Matric Number': result.enrollment.student.matric_number,
                'Student Name': result.enrollment.student.user.get_full_name(),
                'Course Code': result.enrollment.course.code,
                'Course Title': result.enrollment.course.title,
                'Credit Units': result.enrollment.course.credit_units,
                'CA Score': result.ca_score,
                'Exam Score': result.exam_score,
                'Total Score': result.total_score,
                'Grade': result.grade,
                'Grade Point': result.grade_point,
                'Carryover': 'Yes' if result.is_carry_over else 'No',
                'Session': result.enrollment.session.name,
                'Faculty': result.enrollment.student.faculty.name,
                'Department': result.enrollment.student.department.name,
                'Level': result.enrollment.student.current_level.name,
            })
        
        return results_data
    
    @staticmethod
    def get_carryover_data(faculty=None, level=None, session=None):
        """
        Get carryover students data for reporting
        
        Args:
            faculty: Faculty object (optional)
            level: Level object (optional)
            session: AcademicSession object (optional)
        
        Returns:
            List of dictionaries with carryover data
        """
        carryovers = CarryOverList.objects.select_related(
            'result__enrollment__student',
            'result__enrollment__course',
            'session'
        )
        
        # Apply filters
        if faculty:
            carryovers = carryovers.filter(faculty=faculty)
        if level:
            carryovers = carryovers.filter(result__enrollment__student__current_level=level)
        if session:
            carryovers = carryovers.filter(session=session)
        
        carryover_data = []
        for carryover in carryovers:
            carryover_data.append({
                'Matric Number': carryover.result.enrollment.student.matric_number,
                'Student Name': carryover.result.enrollment.student.user.get_full_name(),
                'Course Code': carryover.result.enrollment.course.code,
                'Course Title': carryover.result.enrollment.course.title,
                'Credit Units': carryover.result.enrollment.course.credit_units,
                'Total Score': carryover.result.total_score,
                'Grade': carryover.result.grade,
                'Session': carryover.session.name,
                'Faculty': carryover.faculty.name,
                'Department': carryover.department.name,
                'Level': carryover.result.enrollment.student.current_level.name,
                'Created Date': carryover.created_at.strftime('%Y-%m-%d'),
            })
        
        return carryover_data
    
    @staticmethod
    def get_course_performance_data(faculty=None, department=None, level=None):
        """
        Get course performance data for reporting
        
        Args:
            faculty: Faculty object (optional)
            department: Department object (optional)
            level: Level object (optional)
        
        Returns:
            List of dictionaries with course performance data
        """
        courses = Course.objects.all()
        
        # Apply filters
        if faculty:
            courses = courses.filter(departments__faculty=faculty)
        if department:
            courses = courses.filter(departments=department)
        if level:
            courses = courses.filter(level=level)
        
        performance_data = []
        for course in courses.distinct():
            # Get results for this course
            course_results = Result.objects.filter(
                enrollment__course=course,
                status='PUBLISHED'
            )
            
            total_students = course_results.count()
            if total_students > 0:
                carryovers = course_results.filter(is_carry_over=True).count()
                pass_count = total_students - carryovers
                pass_rate = (pass_count / total_students) * 100
                avg_score = course_results.aggregate(avg_score=Avg('total_score'))['avg_score'] or 0
                
                performance_data.append({
                    'Course Code': course.code,
                    'Course Title': course.title,
                    'Credit Units': course.credit_units,
                    'Level': course.level.name,
                    'Total Students': total_students,
                    'Passed': pass_count,
                    'Carryovers': carryovers,
                    'Pass Rate (%)': round(pass_rate, 2),
                    'Average Score': round(avg_score, 2),
                    'Faculty': ', '.join([dept.faculty.name for dept in course.departments.all()]),
                    'Department': ', '.join([dept.name for dept in course.departments.all()]),
                })
        
        return performance_data
    
    @staticmethod
    def get_student_list_data(faculty=None, department=None, level=None, session=None):
        """
        Get student list data for reporting
        
        Args:
            faculty: Faculty object (optional)
            department: Department object (optional)
            level: Level object (optional)
            session: AcademicSession object (optional)
        
        Returns:
            List of dictionaries with student data
        """
        students = Student.objects.select_related('user', 'faculty', 'department', 'current_level')
        
        # Apply filters
        if faculty:
            students = students.filter(faculty=faculty)
        if department:
            students = students.filter(department=department)
        if level:
            students = students.filter(current_level=level)
        if session:
            students = students.filter(session=session)
        
        student_data = []
        for student in students:
            # Calculate CGPA
            published_results = Result.objects.filter(
                enrollment__student=student,
                status='PUBLISHED'
            )
            
            total_grade_points = sum(
                result.grade_point * result.enrollment.course.credit_units 
                for result in published_results if result.grade_point
            )
            total_credit_units = sum(
                result.enrollment.course.credit_units 
                for result in published_results
            )
            cgpa = total_grade_points / total_credit_units if total_credit_units > 0 else 0.0
            
            student_data.append({
                'Matric Number': student.matric_number,
                'Full Name': student.user.get_full_name(),
                'Email': student.user.email,
                'Faculty': student.faculty.name,
                'Department': student.department.name,
                'Current Level': student.current_level.name,
                'Session': student.session.name if student.session else 'N/A',
                'CGPA': round(cgpa, 2),
                'Total Courses': published_results.count(),
                'Carryovers': published_results.filter(is_carry_over=True).count(),
                'Registration Date': student.created_at.strftime('%Y-%m-%d'),
            })
        
        return student_data


# Utility functions for common report types
def export_faculty_summary(faculty=None):
    """Export faculty summary as Excel"""
    data = ReportingService.get_faculty_summary_data(faculty)
    filename = f"faculty_summary_{faculty.name if faculty else 'all'}"
    return ReportingService.generate_excel_response(data, filename, "Faculty Summary")


def export_student_results(student=None, level=None, faculty=None, session=None):
    """Export student results as Excel"""
    data = ReportingService.get_student_results_data(student, level, faculty, session)
    filename = f"student_results_{session.name if session else 'all'}"
    return ReportingService.generate_excel_response(data, filename, "Student Results")


def export_carryover_list(faculty=None, level=None, session=None):
    """Export carryover list as Excel"""
    data = ReportingService.get_carryover_data(faculty, level, session)
    filename = f"carryover_list_{session.name if session else 'all'}"
    return ReportingService.generate_excel_response(data, filename, "Carryover List")
